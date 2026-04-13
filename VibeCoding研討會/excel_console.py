#!/usr/bin/env python3
"""Console tool to create/fill an Excel file with sample data.

Usage:
  - Provide output path as CLI arg or input when prompted.
  - Example: python excel_console.py output.xlsx
"""
import argparse
import os
import random
import sys

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
except Exception as e:
    print("Missing dependency: openpyxl is required. Install with: pip install openpyxl")
    raise


NAMES = ["王小明", "陳美麗", "李大華", "張婷婷", "林志強", "劉雅婷", "黃子軒"]
JOBS = ["工程師", "老師", "設計師", "醫生", "會計師", "業務", "研究員"]


def generate_rows(n=5):
    rows = []
    for _ in range(n):
        name = random.choice(NAMES)
        age = random.randint(22, 60)
        job = random.choice(JOBS)
        rows.append([name, age, job])
    return rows


def create_or_update(path: str):
    path = os.path.abspath(path)
    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()

    sheet_name = "People"
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)

    headers = ["名字", "年齡", "職業"]
    ws.append(headers)

    rows = generate_rows(5)
    for r in rows:
        ws.append(r)

    last_row = 1 + len(rows)
    table_ref = f"A1:C{last_row}"
    tab = Table(displayName="PeopleTable", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # If default workbook had only the default sheet and it's empty, remove it
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            default = wb["Sheet"]
            if default.max_row == 1 and default.max_column == 1 and default["A1"].value is None:
                wb.remove(default)
        except Exception:
            pass

    wb.save(path)
    return path


def main():
    parser = argparse.ArgumentParser(description="Create an Excel file with sample People data.")
    parser.add_argument("path", nargs="?", help="Output Excel file path (e.g. people.xlsx)")
    args = parser.parse_args()

    out_path = args.path
    if not out_path:
        out_path = input("Enter Excel file path to create (e.g. people.xlsx): ").strip()
    if not out_path:
        print("No path provided. Exiting.")
        sys.exit(1)

    if not out_path.lower().endswith(".xlsx"):
        out_path += ".xlsx"

    try:
        saved = create_or_update(out_path)
        print(f"Excel file created/updated: {saved}")
    except Exception as e:
        print("Failed to create or update Excel file:", e)
        sys.exit(2)


if __name__ == "__main__":
    main()
